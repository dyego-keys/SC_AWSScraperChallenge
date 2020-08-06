import os
import re
from decimal import Decimal
from datetime import datetime
import boto3
#import botocore
from botocore.exceptions import ClientError
from dynamodb_json import json_util
from openpyxl import load_workbook

SHEET_NAME = str(os.environ['SHEET_NAME'])
RETAIL_TABLE = str(os.environ['RETAIL_TABLE'])
RETAIL_INFO_TABLE = str(os.environ['RETAIL_INFO_TABLE'])
REGEX = "\\d{2}-\\w{3}-\\d{4}"


print('** Loading function **')

s3 = boto3.client('s3')
dynamodb = boto3.client('dynamodb')

def parse_xlsx(event, context):
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = event['Records'][0]['s3']['object']['key']

    # Was already processed?
    if was_already_processed(key):
        return "INFO: {} was already processed... nothing to do here...".format(key)
    
    last_report_date = get_last_report_date()
    
    # Parse xlsx file to data
    data = parse_xlsx_to_data(bucket, key, last_report_date)

    # Save items
    save_retail_data(data)

    # Save meta info
    save_retail_info_details(key)

    return "All tasks done!"
    
def was_already_processed(key):
    print("Checking if it was already processed before")
    try:
        response = dynamodb.get_item(
            Key={
                'file_name': {
                    'S':key,
                },
            },
            TableName=RETAIL_INFO_TABLE,
        )
        if 'Item' in response.keys():
            print('INFO: File was already processed!')
            return True
        return False
    except ClientError as client_exception:
        print(client_exception)
    

def get_last_report_date():
    print("Getting the last report date")
    default_date = '19700101'
    try:
        response = dynamodb.scan(
            TableName=RETAIL_INFO_TABLE
        )
        if 'Items' in response and response['Items']:
            result = json_util.loads(response)
            items = result['Items']
            item = sorted(items, key=lambda i: i['processed_at'], reverse=False)[0]
            if 'report_date' in item:
                return item['report_date']
        
        return default_date
        
    except ClientError as client_exception:
        print(client_exception)

def parse_xlsx_to_data(bucket, key, report_date):
    file = get_s3_file(bucket, key)

    print("Parsing file {} filtering by date greater than {}".format(key, report_date))
    data = parse_sheet(file, report_date)
    return data
    
def parse_sheet(file, report_date):

    book = load_workbook(file)
    sheet = book[SHEET_NAME]

    #dates = [(sheet.cell(1, col_index).value).strftime("%Y%m%d") for col_index in range(2, sheet.max_column)]
    dates = []
    col_index = 2
    while True:
        cell_value = sheet.cell(1, col_index).value
        if (cell_value is None):
            break
        dates.append(cell_value.strftime("%Y%m%d"))
        col_index += 1

    row = 2
    data = []
    while True:
        location = sheet.cell(row, 1).value
        if location is None:
            break
        for col in range(1, sheet.max_column - 1):
            if dates[col - 1] > report_date:
                d = {
                    'loc_retail':location,
                    'date_retail':dates[col - 1],
                    'kwh': Decimal(str(sheet.cell(row, col + 1).value))
                }
                data.append(d)
        row = row + 1
    return data

def get_s3_file(bucket, key):
    
    print("Download file from S3 - File name: "+key)
    local_file_name = '/tmp/'+key    
    s3.download_file(bucket, key, local_file_name)
    
    return local_file_name

def save_retail_data(retails):
    
    print("Saving items to dynamoDB")

    for retail in retails:
        response = dynamodb.update_item(
            TableName=RETAIL_TABLE,
            ExpressionAttributeNames={
                '#KWH': 'kwh',
            },
            ExpressionAttributeValues={
                ':kwh': {
                    'N':str(retail['kwh']),
                },
            },
            Key={
                'loc_retail': {
                    'S':retail['loc_retail'],
                },
                'date_retail': {
                    'S':retail['date_retail']
                },
            },
            UpdateExpression='SET #KWH = :kwh',
        )

def save_retail_info_details(key):
    
    print("Saving info data for {} in table {}".format(key, RETAIL_INFO_TABLE))

    match = re.search(REGEX, str(key))
    report_date_obj = datetime.strptime(match.group(0), "%d-%b-%Y")
    report_date = report_date_obj.strftime("%Y%m%d")

    response = dynamodb.put_item(
        TableName=RETAIL_INFO_TABLE,
        Item={
            'file_name': {
                'S':str(key),
            },
            'processed_at': {
                'S':datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
            'report_date': {
                'S':report_date,
            },
        }
    )
    print(response)
    return response